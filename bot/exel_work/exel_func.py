import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import Workbook


file_path = 'C:\\pythonProject\\Main_Diplom\\excel_files\\'
fill = PatternFill(fill_type=None)
thins = Side(border_style="medium", color="000000")


async def create_excel_file(group_number):

    wb = openpyxl.Workbook()
    list = wb.active
    cell = list['B1': 'G1']
    for tuple in cell:
        for el, lab_number in zip(tuple, range(1, 7)):
            el.value = 'Л/р' + f"{lab_number}"
    wb.save(file_path + f"{group_number}.xlsx")


async def add_in_excel(state):
    async with state.proxy() as data:

        wb = load_workbook(file_path + f"{data['group']}.xlsx")
        sheet_group = wb.active
        cell_name = sheet_group['A2': f'A30']
        for tuple in cell_name:
            if tuple[0].value == None:
                tuple[0].value = data['student_name']
                cell_lab = sheet_group[f'B{tuple[0].row}': f'G{tuple[0].row}']
                for list in cell_lab:
                    for param in list:
                        sheet_group[f"{param.coordinate}"].fill = PatternFill('solid', fgColor="FF0000")
                        sheet_group[f"{param.coordinate}"].border = Border(top=thins, bottom=thins,
                                                                           left=thins, right=thins)
                        sheet_group[f"{param.coordinate}"].value = 0
                wb.save(file_path + f"{data['group']}.xlsx")
                return 0


async def visual_edit_excel(state, result):

    async with state.proxy() as data:

        wb = load_workbook(file_path + f"{data['group']}.xlsx")
        sheet_group = wb.active
        cell_name = sheet_group['A2': f'A30']
        for tuple in cell_name:
            if tuple[0].value == data['student_name']:
                coordin = tuple[0].offset(column=int(data['number_lab'])).coordinate
                if result == "unsuccessful":
                    sheet_group[f"{coordin}"].fill = PatternFill('solid', fgColor="ffff00")
                    sheet_group[f"{coordin}"].value += 1
                else:
                    sheet_group[f"{coordin}"].fill = PatternFill('solid', fgColor="008000")
                    sheet_group[f"{coordin}"].value += 1
        wb.save(file_path + f"{data['group']}.xlsx")
        return 0


async def delete_row_excel(state):

    async with state.proxy() as data:

        wb = load_workbook(file_path + f"{data['group']}.xlsx")
        sheet_group = wb.active
        cell_name = sheet_group['A2': f'A30']
        for tuple in cell_name:
            if tuple[0].value == data['student_name']:
                coordin = tuple[0].row
                sheet_group.delete_rows(coordin, 1)
                wb.save(file_path + f"{data['group']}.xlsx")
                return 0




# async def delete_in_excel(state):
#     async with state.proxy() as data:
#
#         wb = load_workbook(file_path + f"{data['group']}.xlsx")
#         sheet_group = wb.active
#         cell_name = sheet_group['A2': f'A30']
#         for tuple in cell_name:
#             if tuple[0].value == data['student_name']:
#                 sheet_group.delete_rows(sheet_group[f"{tuple[0].coordinate}"])
#                 wb.save(file_path + f"{data['group']}.xlsx")
#                 return 0